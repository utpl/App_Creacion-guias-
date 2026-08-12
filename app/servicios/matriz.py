"""Lectura y validacion de la matriz de planificacion.

La columna Unidad/Contenido es la unica fuente autorizada para crear
encabezados tematicos, asi que su integridad es critica.
"""

import csv
import io
import re
import unicodedata
from dataclasses import dataclass, field

from openpyxl import load_workbook

SEMANAS_PERMITIDAS = {8, 16}
MAXIMO_BYTES = 5 * 1024 * 1024
MAXIMO_FILAS = 60

# Variantes aceptadas para cada columna, normalizadas.
EQUIVALENCIAS = {
    "semana": {"semana", "no semana", "num semana", "numero de semana", "n"},
    "resultado_aprendizaje": {
        "resultado de aprendizaje", "resultados de aprendizaje",
        "resultado aprendizaje", "ra",
    },
    "unidad_contenido": {
        "unidad contenido", "unidad/contenido", "unidad contenidos",
        "unidad y contenido", "contenido", "unidad",
    },
    "metodologia": {"metodologia", "metodologias", "metodologia activa"},
    "actividades": {
        "actividades", "actividad", "actividades de aprendizaje",
        "actividades aprendizaje",
    },
}

OBLIGATORIAS = {"semana", "unidad_contenido"}


@dataclass
class FilaLeida:
    semana: int
    resultado_aprendizaje: str
    unidad_contenido: str
    metodologia: str
    actividades: str


@dataclass
class ResultadoMatriz:
    filas: list[FilaLeida] = field(default_factory=list)
    errores: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    @property
    def valida(self) -> bool:
        return not self.errores and bool(self.filas)

    @property
    def semanas_totales(self) -> int:
        return len(self.filas)


def _normalizar(texto) -> str:
    if texto is None:
        return ""
    sin_acentos = "".join(
        c for c in unicodedata.normalize("NFD", str(texto))
        if unicodedata.category(c) != "Mn"
    )
    limpio = re.sub(r"[^a-z0-9 ]", " ", sin_acentos.lower())
    return " ".join(limpio.split())


def _mapear_columnas(encabezados: list[str]) -> dict[str, int]:
    mapa: dict[str, int] = {}
    for posicion, crudo in enumerate(encabezados):
        clave = _normalizar(crudo)
        if not clave:
            continue
        for campo, variantes in EQUIVALENCIAS.items():
            if campo in mapa:
                continue
            if clave in variantes or any(clave.startswith(v) for v in variantes):
                mapa[campo] = posicion
                break
    return mapa


def _leer_tabla(nombre: str, contenido: bytes) -> list[list]:
    if len(contenido) > MAXIMO_BYTES:
        raise ValueError("El archivo supera los 5 MB permitidos.")

    if nombre.lower().endswith(".csv"):
        texto = contenido.decode("utf-8-sig", errors="replace")
        muestra = texto[:2000]
        # Excel en espanol guarda con punto y coma.
        delimitador = ";" if muestra.count(";") > muestra.count(",") else ","
        return [f for f in csv.reader(io.StringIO(texto), delimiter=delimitador)]

    if nombre.lower().endswith((".xlsx", ".xlsm")):
        libro = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
        hoja = libro[libro.sheetnames[0]]
        filas = [list(v) for v in hoja.iter_rows(values_only=True)]
        libro.close()
        return filas

    raise ValueError("Formato no admitido. Use .xlsx o .csv")


def validar(nombre: str, contenido: bytes, semanas_esperadas: int) -> ResultadoMatriz:
    resultado = ResultadoMatriz()

    if semanas_esperadas not in SEMANAS_PERMITIDAS:
        resultado.errores.append("El número de semanas debe ser 8 o 16.")
        return resultado

    try:
        tabla = _leer_tabla(nombre, contenido)
    except ValueError as error:
        resultado.errores.append(str(error))
        return resultado
    except Exception:
        resultado.errores.append("No fue posible leer el archivo.")
        return resultado

    if not tabla:
        resultado.errores.append("El archivo está vacío.")
        return resultado

    if len(tabla) > MAXIMO_FILAS:
        resultado.errores.append(
            f"El archivo tiene más de {MAXIMO_FILAS} filas."
        )
        return resultado

    mapa = _mapear_columnas([str(c) if c is not None else "" for c in tabla[0]])
    faltantes = OBLIGATORIAS - set(mapa)
    if faltantes:
        nombres = {"semana": "Semana", "unidad_contenido": "Unidad/Contenido"}
        resultado.errores.append(
            "No se encontraron las columnas: "
            + ", ".join(nombres[f] for f in sorted(faltantes))
        )
        return resultado

    def celda(fila, campo):
        pos = mapa.get(campo)
        if pos is None or pos >= len(fila) or fila[pos] is None:
            return ""
        return str(fila[pos]).strip()

    vistas: set[int] = set()

    for numero_fila, fila in enumerate(tabla[1:], start=2):
        if all(c is None or str(c).strip() == "" for c in fila):
            continue

        crudo_semana = celda(fila, "semana")
        digitos = re.sub(r"[^0-9]", "", crudo_semana)
        if not digitos:
            resultado.errores.append(
                f"Fila {numero_fila}: la semana no es un número."
            )
            continue

        semana = int(digitos)
        if semana in vistas:
            resultado.errores.append(f"Fila {numero_fila}: semana {semana} repetida.")
            continue
        vistas.add(semana)

        unidad = celda(fila, "unidad_contenido")
        if not unidad:
            resultado.errores.append(
                f"Fila {numero_fila}: Unidad/Contenido está vacía. "
                "Es la única fuente de los temas y no puede faltar."
            )
            continue

        resultado.filas.append(FilaLeida(
            semana=semana,
            resultado_aprendizaje=celda(fila, "resultado_aprendizaje"),
            unidad_contenido=unidad,
            metodologia=celda(fila, "metodologia"),
            actividades=celda(fila, "actividades"),
        ))

    if not resultado.filas:
        resultado.errores.append("No se encontró ninguna fila con datos.")
        return resultado

    resultado.filas.sort(key=lambda f: f.semana)

    # La secuencia debe ir de 1 a N sin huecos.
    esperada = list(range(1, len(resultado.filas) + 1))
    obtenida = [f.semana for f in resultado.filas]
    if obtenida != esperada:
        faltan = sorted(set(esperada) - set(obtenida))
        resultado.errores.append(
            "Las semanas deben ir de 1 a "
            f"{len(resultado.filas)} sin saltos. "
            + (f"Faltan: {', '.join(map(str, faltan))}." if faltan
               else f"Se encontró: {', '.join(map(str, obtenida))}.")
        )
        return resultado

    if len(resultado.filas) != semanas_esperadas:
        resultado.errores.append(
            f"Seleccionó {semanas_esperadas} semanas, pero la matriz "
            f"tiene {len(resultado.filas)}."
        )

    if "metodologia" not in mapa:
        resultado.avisos.append(
            "No se encontró la columna Metodología. La generación no podrá "
            "apoyarse en ella."
        )

    return resultado
