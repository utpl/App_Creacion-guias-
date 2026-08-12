"""Lectura y validacion de archivos de importacion.

Nada se escribe en la base hasta que el usuario confirma la vista previa.
"""

import csv
import io
import re
from dataclasses import dataclass, field

from openpyxl import load_workbook

DOMINIO_INSTITUCIONAL = "@utpl.edu.ec"
PATRON_CORREO = re.compile(r"^[^@\s]+@[^@\s]+\.[a-zA-Z]{2,}$")

# Limites que protegen contra archivos maliciosos o accidentales.
MAXIMO_FILAS = 2000
MAXIMO_BYTES = 5 * 1024 * 1024


@dataclass
class FilaImportada:
    numero: int
    datos: dict[str, str]
    errores: list[str] = field(default_factory=list)
    avisos: list[str] = field(default_factory=list)

    @property
    def valida(self) -> bool:
        return not self.errores


@dataclass
class ResultadoLectura:
    filas: list[FilaImportada]
    error_global: str | None = None

    @property
    def validas(self) -> list[FilaImportada]:
        return [f for f in self.filas if f.valida]

    @property
    def invalidas(self) -> list[FilaImportada]:
        return [f for f in self.filas if not f.valida]


def _normalizar_encabezado(texto: str) -> str:
    return str(texto or "").strip().lower().replace(" ", "_")


def leer_archivo(nombre: str, contenido: bytes) -> list[dict[str, str]]:
    """Convierte CSV o XLSX en una lista de diccionarios."""
    if len(contenido) > MAXIMO_BYTES:
        raise ValueError("El archivo supera los 5 MB permitidos.")

    if nombre.lower().endswith(".csv"):
        texto = contenido.decode("utf-8-sig", errors="replace")
        lector = csv.DictReader(io.StringIO(texto))
        return [
            {_normalizar_encabezado(k): (v or "").strip()
             for k, v in fila.items() if k}
            for fila in lector
        ]

    if nombre.lower().endswith((".xlsx", ".xlsm")):
        libro = load_workbook(
            io.BytesIO(contenido), read_only=True, data_only=True
        )
        hoja = libro[libro.sheetnames[0]]
        iterador = hoja.iter_rows(values_only=True)

        try:
            encabezados = [_normalizar_encabezado(c) for c in next(iterador)]
        except StopIteration:
            raise ValueError("El archivo no contiene datos.")

        filas = []
        for valores in iterador:
            if all(v is None or str(v).strip() == "" for v in valores):
                continue
            filas.append({
                encabezados[i]: str(valores[i]).strip()
                if i < len(valores) and valores[i] is not None else ""
                for i in range(len(encabezados))
            })
        libro.close()
        return filas

    raise ValueError("Formato no admitido. Use .csv o .xlsx")


def validar_usuarios(
    filas_crudas: list[dict[str, str]],
    roles_validos: set[str],
    correos_existentes: set[str],
) -> ResultadoLectura:
    """Valida filas de usuarios. No escribe nada."""

    if not filas_crudas:
        return ResultadoLectura(filas=[], error_global="El archivo no tiene filas.")

    if len(filas_crudas) > MAXIMO_FILAS:
        return ResultadoLectura(
            filas=[],
            error_global=f"El archivo excede las {MAXIMO_FILAS} filas permitidas.",
        )

    requeridas = {"correo", "nombres", "apellidos"}
    presentes = set(filas_crudas[0].keys())
    faltantes = requeridas - presentes
    if faltantes:
        return ResultadoLectura(
            filas=[],
            error_global=(
                "Faltan columnas obligatorias: " + ", ".join(sorted(faltantes))
                + ". Se esperan: correo, nombres, apellidos, roles."
            ),
        )

    resultado: list[FilaImportada] = []
    vistos_en_archivo: set[str] = set()

    for indice, cruda in enumerate(filas_crudas, start=2):
        fila = FilaImportada(numero=indice, datos=dict(cruda))
        correo = cruda.get("correo", "").strip().lower()
        fila.datos["correo"] = correo

        if not correo:
            fila.errores.append("Falta el correo.")
        elif not PATRON_CORREO.match(correo):
            fila.errores.append("El correo no tiene un formato válido.")
        elif not correo.endswith(DOMINIO_INSTITUCIONAL):
            fila.errores.append(
                f"Correo externo. Las cuentas fuera de {DOMINIO_INSTITUCIONAL} "
                "requieren alta manual como invitado."
            )
        elif correo in vistos_en_archivo:
            fila.errores.append("Correo repetido dentro del archivo.")
        elif correo in correos_existentes:
            fila.errores.append("Ya existe una cuenta con este correo.")
        else:
            vistos_en_archivo.add(correo)

        if not cruda.get("nombres", "").strip():
            fila.errores.append("Faltan los nombres.")
        if not cruda.get("apellidos", "").strip():
            fila.errores.append("Faltan los apellidos.")

        crudos = cruda.get("roles", "").strip()
        roles = [r.strip().upper() for r in crudos.split(",") if r.strip()]
        if not roles:
            roles = ["PROFESOR"]
            fila.avisos.append("Sin rol indicado. Se asignará PROFESOR.")
        desconocidos = [r for r in roles if r not in roles_validos]
        if desconocidos:
            fila.errores.append("Roles inexistentes: " + ", ".join(desconocidos))
        fila.datos["roles_procesados"] = ",".join(roles)

        resultado.append(fila)

    return ResultadoLectura(filas=resultado)
