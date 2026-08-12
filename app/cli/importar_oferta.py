"""Importa el catalogo academico desde la hoja de oferta.

Es idempotente: se puede ejecutar varias veces sin duplicar.
Normaliza los valores inconsistentes detectados en el archivo real.
"""

import sys
import unicodedata

from openpyxl import load_workbook

from app.base_datos import FabricaSesion
from app.modelos import Asignatura, Carrera, Facultad

# Correcciones de valores que llegan con variantes en el archivo.
NORMALIZAR_CARRERA = {
    "pedagogia de los idiomas nacionales y extranjeros":
        "Pedagogía de los Idiomas Nacionales y Extranjeros",
}

NORMALIZAR_CAMPO = {
    "integracion": "INTEGRACIÓN",
    "integración": "INTEGRACIÓN",
    "metodolgia": "METODOLOGÍA",
    "metodologia": "METODOLOGÍA",
    "metodología": "METODOLOGÍA",
    "fundamentos teoricos": "FUNDAMENTOS TEÓRICOS",
    "fundamentos teóricos": "FUNDAMENTOS TEÓRICOS",
    "praxis profesional": "PRAXIS PROFESIONAL",
}


def _limpiar(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _clave(texto: str) -> str:
    """Normaliza para comparar: sin acentos, minusculas, sin espacios dobles."""
    sin_acentos = "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )
    return " ".join(sin_acentos.lower().split())


def _titulo(texto: str) -> str:
    """Capitaliza respetando preposiciones."""
    menores = {"de", "del", "la", "las", "los", "y", "en", "para", "e"}
    palabras = texto.split()
    return " ".join(
        p.capitalize() if i == 0 or p.lower() not in menores else p.lower()
        for i, p in enumerate(palabras)
    )


def importar(ruta: str) -> None:
    libro = load_workbook(ruta, read_only=True, data_only=True)
    hoja = libro[libro.sheetnames[0]]
    filas = hoja.iter_rows(values_only=True)

    encabezados = [_limpiar(c) for c in next(filas)]
    indice = {nombre: i for i, nombre in enumerate(encabezados)}

    requeridas = [
        "Código Banner", "Nombre de Asignatura", "Facultad",
        "Carrera responsable de la asignatura",
    ]
    faltantes = [c for c in requeridas if c not in indice]
    if faltantes:
        sys.exit("Faltan columnas en el archivo: " + ", ".join(faltantes))

    def valor(fila, columna):
        pos = indice.get(columna)
        return _limpiar(fila[pos]) if pos is not None and pos < len(fila) else ""

    with FabricaSesion() as bd:
        facultades: dict[str, Facultad] = {}
        carreras: dict[tuple[str, str], Carrera] = {}
        nuevas_asignaturas = 0
        ya_existentes = 0

        for fila in filas:
            codigo = valor(fila, "Código Banner")
            if not codigo:
                continue

            # --- Facultad ---
            nombre_facultad = _titulo(valor(fila, "Facultad"))
            clave_f = _clave(nombre_facultad)
            if clave_f not in facultades:
                existente = bd.query(Facultad).filter(
                    Facultad.nombre == nombre_facultad
                ).one_or_none()
                if existente is None:
                    existente = Facultad(nombre=nombre_facultad)
                    bd.add(existente)
                    bd.flush()
                facultades[clave_f] = existente
            facultad = facultades[clave_f]

            # --- Carrera ---
            cruda = valor(fila, "Carrera responsable de la asignatura")
            # Una fila trae dos carreras separadas por salto de linea.
            cruda = cruda.split("\n")[0].strip()
            clave_c = _clave(cruda)
            nombre_carrera = NORMALIZAR_CARRERA.get(clave_c, _titulo(cruda))

            llave = (clave_f, _clave(nombre_carrera))
            if llave not in carreras:
                existente = bd.query(Carrera).filter(
                    Carrera.facultad_id == facultad.id,
                    Carrera.nombre == nombre_carrera,
                ).one_or_none()
                if existente is None:
                    existente = Carrera(
                        facultad_id=facultad.id,
                        nombre=nombre_carrera,
                        departamento=_titulo(valor(fila, "Departamento")) or None,
                        nivel="GRADO",
                        modalidad="EN_LINEA",
                    )
                    bd.add(existente)
                    bd.flush()
                carreras[llave] = existente
            carrera = carreras[llave]

            # --- Asignatura ---
            if bd.query(Asignatura).filter(Asignatura.codigo == codigo).first():
                ya_existentes += 1
                continue

            campo_crudo = valor(fila, "Tipo de componenteo asignatura / CAMPO DE FORMACIÓN")
            campo = NORMALIZAR_CAMPO.get(_clave(campo_crudo), campo_crudo or None)

            creditos_crudo = valor(fila, "Créditos")
            try:
                creditos = int(float(creditos_crudo)) if creditos_crudo else None
            except ValueError:
                creditos = None

            bd.add(Asignatura(
                codigo=codigo,
                nombre=valor(fila, "Nombre de Asignatura"),
                carrera_id=carrera.id,
                creditos=creditos,
                ciclo=valor(fila, "Ciclo único") or valor(fila, "Ciclos") or None,
                campo_formacion=campo,
                url_canvas=valor(fila, "URL CANVAS") or None,
            ))
            nuevas_asignaturas += 1

        bd.commit()
        libro.close()

        print(f"Facultades:  {len(facultades)}")
        print(f"Carreras:    {len(carreras)}")
        print(f"Asignaturas: {nuevas_asignaturas} nuevas, {ya_existentes} ya existían")


if __name__ == "__main__":
    if len(sys.argv) != 2:
        sys.exit("Uso: python -m app.cli.importar_oferta datos_ejemplo/oferta.xlsx")
    importar(sys.argv[1])
