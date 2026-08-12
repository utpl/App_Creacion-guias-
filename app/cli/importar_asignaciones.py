"""Importa periodo, docentes y asignaciones desde la hoja de oferta.

Reglas aplicadas:
  - La cadena de 'Autor GD' puede traer varias personas separadas por saltos.
  - El primero es titular; el resto, colaboradores.
  - Se deduplica por nombre dentro de la misma cadena.
  - Los correos fuera de @utpl.edu.ec NO crean cuenta: quedan reportados.
  - Las cuentas nacen sin contrasena.
Es idempotente.
"""

import sys

from openpyxl import load_workbook

from app.base_datos import FabricaSesion
from app.modelos import (
    AsignacionDocente,
    Asignatura,
    PeriodoAcademico,
    Rol,
    Usuario,
    UsuarioRol,
)

DOMINIO = "@utpl.edu.ec"
ETIQUETA_REESTRUCTURA = "reestructurada por"


def _limpiar(valor) -> str:
    return str(valor).strip() if valor is not None else ""


def _partes(celda: str) -> list[str]:
    """Separa una celda multilinea y descarta la etiqueta de reestructura."""
    return [
        p.strip()
        for p in celda.split("\n")
        if p.strip() and ETIQUETA_REESTRUCTURA not in p.strip().lower()
    ]


def _dividir_nombre(completo: str) -> tuple[str, str]:
    """Convencion ecuatoriana: los dos ultimos terminos son apellidos."""
    palabras = completo.split()

    if len(palabras) <= 2:
        return (
            palabras[0] if palabras else completo,
            " ".join(palabras[1:]),
        )

    return (
        " ".join(palabras[:-2]),
        " ".join(palabras[-2:]),
    )


def importar(ruta: str, nombre_periodo: str) -> None:
    libro = load_workbook(
        ruta,
        read_only=True,
        data_only=True,
    )

    hoja = libro[libro.sheetnames[0]]
    filas = hoja.iter_rows(values_only=True)

    encabezados = [_limpiar(c) for c in next(filas)]
    indice = {n: i for i, n in enumerate(encabezados)}

    for columna in [
        "Código Banner",
        "Autor GD",
        "Correo electronico GD",
    ]:
        if columna not in indice:
            libro.close()
            sys.exit(
                f"Falta la columna '{columna}' en el archivo."
            )

    def valor(fila, columna):
        pos = indice.get(columna)

        if pos is not None and pos < len(fila):
            return _limpiar(fila[pos])

        return ""

    with FabricaSesion() as bd:
        # --- Periodo ---
        periodo = (
            bd.query(PeriodoAcademico)
            .filter_by(nombre=nombre_periodo)
            .one_or_none()
        )

        if periodo is None:
            bd.query(PeriodoAcademico).update(
                {"activo": False}
            )

            periodo = PeriodoAcademico(
                nombre=nombre_periodo,
                activo=True,
            )

            bd.add(periodo)
            bd.flush()

            print(f"Periodo creado: {nombre_periodo}")
        else:
            print(f"Periodo existente: {nombre_periodo}")

        rol_profesor = (
            bd.query(Rol)
            .filter_by(codigo="PROFESOR")
            .one_or_none()
        )

        if rol_profesor is None:
            libro.close()
            sys.exit(
                "Ejecute primero: "
                "python -m app.cli.sembrar_roles"
            )

        docentes_creados = 0
        vinculos_creados = 0
        vinculos_existentes = 0

        externos: list[tuple[str, str, str]] = []
        desajustes: list[str] = []
        sin_asignatura: list[str] = []

        for fila in filas:
            codigo = valor(fila, "Código Banner")

            if not codigo:
                continue

            asignatura = (
                bd.query(Asignatura)
                .filter_by(codigo=codigo)
                .one_or_none()
            )

            if asignatura is None:
                sin_asignatura.append(codigo)
                continue

            nombres_crudos = _partes(
                valor(fila, "Autor GD")
            )

            correos_crudos = _partes(
                valor(fila, "Correo electronico GD")
            )

            # Deduplicar por nombre conservando el orden.
            vistos: set[str] = set()
            nombres = []

            for nombre in nombres_crudos:
                clave = nombre.lower()

                if clave not in vistos:
                    vistos.add(clave)
                    nombres.append(nombre)

            if len(nombres) != len(correos_crudos):
                desajustes.append(
                    f"{codigo}: "
                    f"{len(nombres)} nombre(s), "
                    f"{len(correos_crudos)} correo(s)"
                )
                continue

            for posicion, (nombre, correo) in enumerate(
                zip(nombres, correos_crudos)
            ):
                correo = correo.lower()

                if not correo.endswith(DOMINIO):
                    externos.append(
                        (codigo, nombre, correo)
                    )
                    continue

                docente = (
                    bd.query(Usuario)
                    .filter_by(correo=correo)
                    .one_or_none()
                )

                if docente is None:
                    pila, apellidos = _dividir_nombre(
                        nombre
                    )

                    docente = Usuario(
                        correo=correo,
                        nombres=pila,
                        apellidos=apellidos,
                        hash_password=None,
                        requiere_cambio_password=True,
                        origen="NOMINA",
                    )

                    bd.add(docente)
                    bd.flush()

                    bd.add(
                        UsuarioRol(
                            usuario_id=docente.id,
                            rol_id=rol_profesor.id,
                        )
                    )

                    docentes_creados += 1

                existente = (
                    bd.query(AsignacionDocente)
                    .filter_by(
                        docente_id=docente.id,
                        asignatura_id=asignatura.id,
                        periodo_id=periodo.id,
                    )
                    .one_or_none()
                )

                if existente:
                    vinculos_existentes += 1
                    continue

                bd.add(
                    AsignacionDocente(
                        docente_id=docente.id,
                        asignatura_id=asignatura.id,
                        periodo_id=periodo.id,
                        rol_en_asignatura=(
                            "titular"
                            if posicion == 0
                            else "colaborador"
                        ),
                    )
                )

                vinculos_creados += 1

        bd.commit()

    libro.close()

    print()
    print(f"Docentes creados:       {docentes_creados}")
    print(f"Vinculos creados:       {vinculos_creados}")
    print(
        f"Vinculos ya existentes: {vinculos_existentes}"
    )

    if externos:
        print(
            f"\nCorreos externos ({len(externos)}) "
            "— requieren alta manual:"
        )

        for codigo, nombre, correo in externos:
            print(
                f"  {codigo}  {nombre}  <{correo}>"
            )

    if desajustes:
        print(
            f"\nFilas con datos incompletos "
            f"({len(desajustes)}):"
        )

        for desajuste in desajustes:
            print(f"  {desajuste}")

    if sin_asignatura:
        print(
            "\nCodigos sin asignatura en el catalogo: "
            + ", ".join(sin_asignatura)
        )


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit(
            "Uso: python -m app.cli.importar_asignaciones "
            'datos_ejemplo/oferta.xlsx '
            '"Oct. 2026 - Feb. 2027"'
        )

    importar(
        sys.argv[1],
        sys.argv[2],
    )

